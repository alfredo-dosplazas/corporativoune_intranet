from django.utils.timezone import now
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from apps.core.models import Empresa
from apps.papeleria.models.requisiciones import Requisicion
from apps.papeleria.services.articulo_acumulado_report import articulo_acumulado_report


def articulo_acumulado_excel(empresa_ids=()):
    data = articulo_acumulado_report(empresa_ids=empresa_ids)
    empresas = Empresa.objects.filter(pk__in=empresa_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Acumulado de Artículos'

    # =====================
    # Estilos base
    # =====================
    bold = Font(bold=True)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    currency = numbers.FORMAT_CURRENCY_USD_SIMPLE
    percent = numbers.FORMAT_PERCENTAGE_00

    # =====================
    # TÍTULO
    # =====================
    ws.merge_cells("A1:I1")
    ws["A1"] = "REPORTE DE ARTÍCULOS ACUMULADOS"
    ws["A1"].font = title_font
    ws["A1"].alignment = center

    ws.merge_cells("A2:I2")
    ws["A2"] = f"Generado el {now().strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].alignment = center

    # =====================
    # DATOS GENERALES
    # =====================
    fila = 4

    datos = [
        ("Empresas:", ", ".join(e.nombre_corto for e in empresas) or "TODAS"),
        # ("Razones Sociales:", "TODAS"),
        ("Estado Requisiciones:", "Autorizadas por Contraloría"),
    ]

    for label, value in datos:
        ws[f"A{fila}"] = label
        ws[f"A{fila}"].font = bold
        ws.merge_cells(f"B{fila}:I{fila}")
        ws[f"B{fila}"] = value
        fila += 1

    fila += 1

    # =====================
    # ENCABEZADOS TABLA
    # =====================
    headers = [
        "Código VS DP",
        "Número Papelería",
        "Artículo",
        "Unidad",
        "Precio",
        "Impuesto",
        "Importe Unitario",
        "Cantidad Autorizada",
        "Subtotal",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila, column=col, value=header)
        cell.font = bold
        cell.alignment = center
        cell.border = thin

    fila_inicio_detalle = fila + 1
    fila += 1

    # =====================
    # DETALLE
    # =====================
    for record in data:
        ws[f"A{fila}"] = record['codigo_vs_dp']
        ws[f"B{fila}"] = record['numero_papeleria']
        ws[f"C{fila}"] = record['articulo']
        ws[f"D{fila}"] = record['unidad']

        ws[f"E{fila}"] = record['precio']
        ws[f"E{fila}"].number_format = currency

        ws[f"F{fila}"] = record['impuesto']
        ws[f"F{fila}"].number_format = percent

        ws[f"G{fila}"] = record['importe_unitario']
        ws[f"G{fila}"].number_format = currency

        ws[f"H{fila}"] = record['cantidad_total_autorizada']
        ws[f"H{fila}"].alignment = right

        ws[f"I{fila}"] = record['importe_total']
        ws[f"I{fila}"].number_format = currency

        for col in range(1, 10):
            ws.cell(row=fila, column=col).border = thin

        fila += 1

    fila_fin_detalle = fila - 1

    # =====================
    # TOTALES
    # =====================
    ws.merge_cells(f"A{fila}:H{fila}")
    ws[f"A{fila}"] = "TOTAL GENERAL"
    ws[f"A{fila}"].font = bold
    ws[f"A{fila}"].alignment = right

    ws[f"I{fila}"] = f"=SUM(I{fila_inicio_detalle}:I{fila_fin_detalle})"
    ws[f"I{fila}"].font = bold
    ws[f"I{fila}"].number_format = currency

    for col in range(1, 10):
        ws.cell(row=fila, column=col).border = thin

    # =====================
    # AJUSTE DE ANCHOS
    # =====================
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # =====================
    # HOJA DE REQUISICIONES
    # =====================
    ws_req = wb.create_sheet('Requisiciones Incluidas')

    ws_req.append([
        "ID Requisición",
        "Folio",
        "Solicitante",
        "Empresa",
        "Fecha Creación",
        "Fecha Autorización",
        "Estado",
    ])

    for cell in ws_req[1]:
        cell.font = bold
        cell.border = thin
        cell.alignment = center

    requisiciones = (
        Requisicion.objects
        .filter(
            estado='autorizada_contraloria',
            empresa_id__in=empresa_ids if empresa_ids else Empresa.objects.values_list('id', flat=True),
        )
        .distinct()
    )

    for r in requisiciones:
        ws_req.append([
            r.id,
            r.folio,
            r.solicitante.contacto.nombre_completo,
            r.empresa.nombre_corto,
            r.created_at.strftime('%d/%m/%Y'),
            r.fecha_autorizacion_contraloria.strftime('%d/%m/%Y') if r.fecha_autorizacion_contraloria else '',
            r.get_estado_display(),
        ])

    # AJUSTE TAMAÑOS
    COLUMN_WIDTHS = {
        "A": 18,  # ID Requisición
        "B": 30,  # Folio
        "C": 30,  # Solicitante
        "D": 12,  # Empresa
        "E": 15,  # Fecha Creación
        "F": 20,  # Fecha Autorización
        "G": 30,  # Estado
    }

    for col, width in COLUMN_WIDTHS.items():
        ws_req.column_dimensions[col].width = width

    return wb
