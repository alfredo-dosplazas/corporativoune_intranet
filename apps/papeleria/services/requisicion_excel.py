import os

from django.conf import settings
from django.utils.timezone import localtime
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.workbook import Workbook

from apps.papeleria.models.requisiciones import Requisicion


def requisicion_excel(requisicion: Requisicion):
    wb = Workbook()
    ws = wb.active
    ws.title = requisicion.folio

    # =====================
    # Estilos base
    # =====================
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    currency = numbers.FORMAT_CURRENCY_USD_SIMPLE
    percent = numbers.FORMAT_PERCENTAGE

    # =====================
    # TÍTULO
    # =====================
    ws.merge_cells("A1:J1")
    ws["A1"] = "REQUISICIÓN DE PAPELERÍA"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:J2")
    ws["A2"] = requisicion.folio
    ws["A2"].alignment = center

    # =====================
    # LOGO
    # =====================
    if requisicion.empresa.logo:
        logo_path = os.path.join(settings.MEDIA_ROOT, requisicion.empresa.logo.name)
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 150
            img.height = 120
            ws.add_image(img, "A1")

    # =====================
    # DATOS GENERALES
    # =====================
    fila = 5

    datos = [
        ("Empresa:", requisicion.empresa.nombre),
        ("Folio:", requisicion.folio),
        ("Solicitante:", getattr(getattr(requisicion.solicitante, 'contacto', requisicion.solicitante.get_full_name()),
                                 'nombre_completo', '')),
        ("Departamento:",
         requisicion.solicitante.profile.departamento if hasattr(requisicion.solicitante, "profile") else ""),
        ("Fecha solicitud:", localtime(requisicion.created_at).strftime("%d/%m/%Y")),
    ]

    for label, value in datos:
        ws[f"A{fila}"] = label
        ws[f"A{fila}"].font = bold
        ws.merge_cells(f"B{fila}:J{fila}")
        ws[f"B{fila}"] = value
        fila += 1

    fila += 1

    # =====================
    # ENCABEZADOS TABLA
    # =====================
    headers = [
        "Código VS DP",
        "Número Papelería",
        "Artículo",
        "Cantidad",
        "Cant. liberada",
        "Unidad",
        "Precio",
        "Impuesto",
        "Importe",
        "Subtotal",
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=fila, column=col, value=header)
        cell.font = bold
        cell.alignment = center
        cell.border = thin

    fila_inicio_detalle = fila + 1
    fila += 1

    # =====================
    # DETALLE (CON FÓRMULAS)
    # =====================
    for detalle in requisicion.detalle_requisicion.all():
        ws[f"A{fila}"] = detalle.articulo.codigo_vs_dp
        ws[f"B{fila}"] = detalle.articulo.numero_papeleria
        ws[f"C{fila}"] = detalle.articulo.nombre
        ws[f"D{fila}"] = detalle.cantidad
        ws[f"E{fila}"] = detalle.cantidad_liberada
        ws[f"F{fila}"] = detalle.articulo.unidad.clave
        ws[f"G{fila}"] = detalle.articulo.precio
        ws[f"H{fila}"] = detalle.articulo.impuesto
        ws[f"I{fila}"] = detalle.articulo.importe

        ws[f"G{fila}"].number_format = currency
        ws[f"H{fila}"].number_format = percent

        # Importe = Impuesto * Precio
        ws[f"I{fila}"] = f"=(1+H{fila})*G{fila}"
        ws[f"I{fila}"].number_format = currency

        # Subtotal = Cantidad liberada * Importe (o cantidad)
        ws[f"J{fila}"] = f"=IF(E{fila}>0,E{fila}*I{fila},D{fila}*I{fila})"
        ws[f"J{fila}"].number_format = currency

        # Bordes
        for col in range(1, 11):
            ws.cell(row=fila, column=col).border = thin

        fila += 1

        # =====================
        # TOTAL
        # =====================
    ws.merge_cells(f"A{fila}:I{fila}")
    ws[f"A{fila}"] = "TOTAL"
    ws[f"A{fila}"].font = bold
    ws[f"A{fila}"].alignment = Alignment(horizontal="right")

    ws[f"J{fila}"] = f"=SUM(J{fila_inicio_detalle}:J{fila - 1})"
    ws[f"J{fila}"].font = bold
    ws[f"J{fila}"].number_format = currency

    fila += 2

    # =====================
    # FIRMAS
    # =====================
    bloques_firma = [
        ("A", "C"),
        ("D", "F"),
        ("G", "I"),
        ("J", "L"),
    ]

    firmas = [
        ("Solicitante", requisicion.solicitante.get_full_name(), requisicion.aprobo_solicitante),
        ("Aprobador", requisicion.aprobador.get_full_name(), requisicion.aprobo_aprobador),
        ("Compras", requisicion.compras.get_full_name(), requisicion.aprobo_compras),
        ("Contraloría", requisicion.contraloria.get_full_name(), requisicion.aprobo_contraloria),
    ]

    ws.row_dimensions[fila].height = 30
    ws.row_dimensions[fila + 1].height = 22
    ws.row_dimensions[fila + 2].height = 22

    for (col_ini, col_fin), (rol, nombre, aprobado) in zip(bloques_firma, firmas):
        ws.merge_cells(f"{col_ini}{fila}:{col_fin}{fila}")
        ws[f"{col_ini}{fila}"] = "______________________________"
        ws[f"{col_ini}{fila}"].alignment = center

        ws.merge_cells(f"{col_ini}{fila + 1}:{col_fin}{fila + 1}")
        ws[f"{col_ini}{fila + 1}"] = nombre
        ws[f"{col_ini}{fila + 1}"].alignment = center
        ws[f"{col_ini}{fila + 1}"].font = bold

        ws.merge_cells(f"{col_ini}{fila + 2}:{col_fin}{fila + 2}")
        ws[f"{col_ini}{fila + 2}"] = (
            f"{rol} – AUTORIZADO DIGITALMENTE"
            if aprobado else
            f"{rol} – PENDIENTE"
        )
        ws[f"{col_ini}{fila + 2}"].alignment = center

    # AJUSTE TAMAÑOS
    COLUMN_WIDTHS = {
        "A": 18,  # Código VS DP
        "B": 20,  # Número Papelería
        "C": 40,  # Artículo
        "D": 12,  # Cantidad
        "E": 15,  # Cant. liberada
        "F": 12,  # Unidad
        "G": 14,  # Precio
        "H": 12,  # Impuesto
        "I": 16,  # Importe
        "J": 18,  # Subtotal
    }

    for col, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 25

    for r in range(5, fila_inicio_detalle):
        ws.row_dimensions[r].height = 22

    return wb
