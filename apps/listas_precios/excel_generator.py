import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def generate_excel_report(output_path, listas, lineas_data, productos_data):
    wb = openpyxl.Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    # Estilos básicos
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    border_thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(
        left=border_thin, right=border_thin, top=border_thin, bottom=border_thin
    )

    # ==========================================
    # HOJA 1: Lineas (Configuración de Descuentos/Utilidades)
    # ==========================================
    ws_lineas = wb.create_sheet(title="Lineas")
    headers_lineas = ["Clave Línea", "Descripción Línea"]
    for l in listas:
        cve_lista = str(l["CVE_PRECIO"])
        nombre_lista = l.get("DESCRIPCION", f"Lista {cve_lista}")
        headers_lineas.extend(
            [f"Desc % ({nombre_lista})", f"Util % ({nombre_lista})"]
        )

    ws_lineas.append(headers_lineas)

    # Mapeo para saber en qué fila está cada línea dentro de la hoja Lineas
    linea_row_map = {}
    row_idx = 2
    for item in lineas_data:
        cve_lin = item["linea"]
        desc_lin = item["desc_lin"]
        row_vals = [cve_lin, desc_lin]

        for l in listas:
            cve_lista = str(l["CVE_PRECIO"])
            conf = item["listas"].get(cve_lista, {"desc": 0.0, "util": 0.0})
            row_vals.extend([conf["desc"] / 100.0, conf["util"] / 100.0])

        ws_lineas.append(row_vals)
        linea_row_map[cve_lin] = row_idx
        row_idx += 1

    # Formato a la Hoja Líneas
    for cell in ws_lineas[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws_lineas.max_row + 1):
        for c in range(3, ws_lineas.max_column + 1):
            ws_lineas.cell(row=r, column=c).number_format = "0.00%"

    # ==========================================
    # HOJA 2: Productos (Catálogo Base)
    # ==========================================
    ws_prod = wb.create_sheet(title="Productos")
    ws_prod.append(
        ["CVE_ART", "DESCRIPCION", "CVE_LIN", "DESC_LIN", "PRECIO_LISTA"]
    )

    prod_row_map = {}
    r_idx = 2
    for p in productos_data:
        cve_art = str(p.get("CVE_ART", ""))
        desc_art = str(p.get("DESCRIPCION", ""))
        cve_lin = str(p.get("CVE_LIN", ""))
        desc_lin = str(p.get("DESC_LIN", ""))
        precio_lista = float(p.get("PRECIO_LISTA", 0.0) or 0.0)

        ws_prod.append([cve_art, desc_art, cve_lin, desc_lin, precio_lista])
        prod_row_map[cve_art] = r_idx
        r_idx += 1

    for cell in ws_prod[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws_prod.max_row + 1):
        ws_prod.cell(row=r, column=5).number_format = "$#,##0.00"

    # ==========================================
    # HOJA 3: Productos_Con_Precios (Consolidado con Fórmulas)
    # ==========================================
    ws_prod_precios = wb.create_sheet(title="Productos_Con_Precios")
    headers_all = [
        "CVE_ART",
        "DESCRIPCION",
        "CVE_LIN",
        "DESC_LIN",
        "PRECIO_LISTA",
    ]

    for l in listas:
        cve_lista = str(l["CVE_PRECIO"])
        nombre_lista = l.get("DESCRIPCION", f"Lista {cve_lista}")
        headers_all.extend(
            [
                f"Desc % ({nombre_lista})",
                f"Util % ({nombre_lista})",
                f"Precio Final ({nombre_lista})",
            ]
        )

    ws_prod_precios.append(headers_all)

    # Fórmula para Precio Final:
    # Precio Final = (PRECIO_LISTA * (1 - Desc%)) * (1 + Util%)
    r_idx = 2
    for p in productos_data:
        cve_art = str(p.get("CVE_ART", ""))
        cve_lin = str(p.get("CVE_LIN", ""))

        row_cells = [
            f"=Productos!A{r_idx}",
            f"=Productos!B{r_idx}",
            f"=Productos!C{r_idx}",
            f"=Productos!D{r_idx}",
            f"=Productos!E{r_idx}",
        ]

        # Calcular posición de columnas según las listas
        # Columna 1=A, 2=B, 3=C, 4=D, 5=E (PRECIO_LISTA)
        # Cada lista agrega 3 columnas (Desc, Util, PrecioFinal)
        col_offset = 6
        for idx_l, l in enumerate(listas):
            # Buscar en la hoja Lineas la clave de línea
            # En Hoja Lineas: A=CVE_LIN. Las cols de Desc y Util varían por lista
            col_desc_lineas = get_column_letter(3 + (idx_l * 2))
            col_util_lineas = get_column_letter(4 + (idx_l * 2))

            # Fórmulas BUSCARV/VLOOKUP a la hoja Lineas usando la línea del producto (=C{r_idx})
            formula_desc = (
                f'=IFERROR(VLOOKUP(C{r_idx}, Lineas!A:ZZ, {3 + (idx_l * 2)}, FALSE), 0)'
            )
            formula_util = (
                f'=IFERROR(VLOOKUP(C{r_idx}, Lineas!A:ZZ, {4 + (idx_l * 2)}, FALSE), 0)'
            )

            col_letter_desc = get_column_letter(col_offset)
            col_letter_util = get_column_letter(col_offset + 1)

            # Fórmula Precio Final = (E{row} * (1 - Desc)) * (1 + Util)
            formula_precio = f'=(E{r_idx} * (1 - {col_letter_desc}{r_idx})) / (1 - {col_letter_util}{r_idx})'

            row_cells.extend([formula_desc, formula_util, formula_precio])
            col_offset += 3

        ws_prod_precios.append(row_cells)
        r_idx += 1

    # Formato Hoja Productos_Con_Precios
    for cell in ws_prod_precios[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(2, ws_prod_precios.max_row + 1):
        ws_prod_precios.cell(row=r, column=5).number_format = "$#,##0.00"
        c_off = 6
        for _ in listas:
            ws_prod_precios.cell(row=r, column=c_off).number_format = "0.00%"
            ws_prod_precios.cell(row=r, column=c_off + 1).number_format = (
                "0.00%"
            )
            ws_prod_precios.cell(row=r, column=c_off + 2).number_format = (
                "$#,##0.00"
            )
            c_off += 3

    # ==========================================
    # HOJA 4: Una Hoja Individual Por Cada Lista de Precio
    # ==========================================
    for idx_l, l in enumerate(listas):
        cve_lista = str(l["CVE_PRECIO"])
        nombre_lista = l.get("DESCRIPCION", f"Lista_{cve_lista}")

        # Limpiar nombre de hoja (Excel no admite caracteres inválidos)
        sheet_title = f"Lista_{cve_lista}_{nombre_lista}"[
            :31
        ]  # Max 31 chars
        for char in ["\\", "/", "?", "*", "[", "]"]:
            sheet_title = sheet_title.replace(char, "")

        ws_list = wb.create_sheet(title=sheet_title)
        ws_list.append(
            [
                "CVE_ART",
                "DESCRIPCION",
                "CVE_LIN",
                "DESC_LIN",
                "PRECIO_LISTA",
                "DESC %",
                "UTIL %",
                f"PRECIO {nombre_lista.upper()}",
            ]
        )

        # Ubicación en la hoja Productos_Con_Precios de las columnas correspondientes
        # Columna de Desc % de esta lista: 6 + (idx_l * 3)
        # Columna de Util % de esta lista: 7 + (idx_l * 3)
        # Columna de Precio Final: 8 + (idx_l * 3)
        col_desc_letter = get_column_letter(6 + (idx_l * 3))
        col_util_letter = get_column_letter(7 + (idx_l * 3))
        col_prec_letter = get_column_letter(8 + (idx_l * 3))

        r_idx = 2
        for _ in productos_data:
            ws_list.append(
                [
                    f"=Productos_Con_Precios!A{r_idx}",
                    f"=Productos_Con_Precios!B{r_idx}",
                    f"=Productos_Con_Precios!C{r_idx}",
                    f"=Productos_Con_Precios!D{r_idx}",
                    f"=Productos_Con_Precios!E{r_idx}",
                    f"=Productos_Con_Precios!{col_desc_letter}{r_idx}",
                    f"=Productos_Con_Precios!{col_util_letter}{r_idx}",
                    f"=Productos_Con_Precios!{col_prec_letter}{r_idx}",
                ]
            )
            r_idx += 1

        # Formato de la Hoja de Lista Individual
        for cell in ws_list[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in range(2, ws_list.max_row + 1):
            ws_list.cell(row=r, column=5).number_format = "$#,##0.00"
            ws_list.cell(row=r, column=6).number_format = "0.00%"
            ws_list.cell(row=r, column=7).number_format = "0.00%"
            ws_list.cell(row=r, column=8).number_format = "$#,##0.00"

    # Autoajustar ancho de columnas para todas las hojas
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Guardar libro
    wb.save(output_path)