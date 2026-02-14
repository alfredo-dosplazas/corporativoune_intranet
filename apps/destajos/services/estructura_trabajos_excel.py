from openpyxl.styles import Font, Alignment, Border, Side, numbers
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import TableStyleInfo, Table

from apps.destajos.models import Estructura


def estructura_trabajos_excel(estructura: Estructura):
    wb = Workbook()
    ws = wb.active
    ws.title = estructura.nombre[:31]  # Excel limita a 31 caracteres

    # =====================
    # Estilos base
    # =====================
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    currency = numbers.FORMAT_CURRENCY_USD_SIMPLE

    trabajos = estructura.trabajos.all()

    encabezados = ['Estructura', 'Trabajo', 'Contratista', 'Precio']
    ws.append(encabezados)

    # Aplicar estilo a encabezados
    for col in range(1, len(encabezados) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = bold
        cell.alignment = center
        cell.border = thin

    # Datos
    for t in trabajos:
        estr = estructura.nombre
        trabajo = t.trabajo.nombre
        contratista = getattr(getattr(t, 'contratista', None), 'nombre', None)
        precio = t.precio_vigente

        ws.append([estr, trabajo, contratista, precio])

        # Aplicar formato moneda a la última celda
        ws.cell(row=ws.max_row, column=4).number_format = currency

    # =====================
    # Crear tabla estilo Excel (Ctrl + T)
    # =====================
    last_row = ws.max_row
    last_col = ws.max_column

    table_range = f"A1:{ws.cell(row=last_row, column=last_col).coordinate}"

    table = Table(
        displayName="TablaEstructuraTrabajos",
        ref=table_range
    )

    style = TableStyleInfo(
        name="TableStyleMedium9",  # puedes cambiar el estilo
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    table.tableStyleInfo = style
    ws.add_table(table)

    return wb
