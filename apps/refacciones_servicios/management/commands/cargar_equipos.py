import os

from django.core.management import BaseCommand
from openpyxl.reader.excel import load_workbook

from apps.refacciones_servicios.models import Equipo


class Command(BaseCommand):
    help = 'Carga el catálogo de equipos desde un archivo Excel'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta del archivo Excel de equipos (.xlsx)')

    def handle(self, *args, **options):
        ruta_archivo = options['excel_file']

        if not os.path.exists(ruta_archivo):
            self.stdout.write(self.style.ERROR(f"El archivo no existe: {ruta_archivo}"))
            return

        self.stdout.write(self.style.SUCCESS(f"Leyendo catálogo de equipos: {ruta_archivo}"))

        # Abrimos el libro. Si tiene una sola hoja, usamos active.
        wb = load_workbook(ruta_archivo, data_only=True)
        ws = wb.active

        # Ajusta la fila donde empiezan los datos reales (ej: fila 2 si la 1 son encabezados)
        fila_inicio = 2
        fila_fin = ws.max_row

        equipos_creados = 0
        equipos_actualizados = 0

        for r in range(fila_inicio, fila_fin + 1):
            # 1. Extraer los valores de las columnas indicadas
            val_planta = ws[f'B{r}'].value  # Ubicación / Planta
            val_operador = ws[f'C{r}'].value  # Operador
            val_equipo = ws[f'D{r}'].value  # Equipo (Nombre o a veces Identificador)
            val_identif = ws[f'E{r}'].value  # Año / Identificador
            val_serie_m = ws[f'F{r}'].value  # Serie / Marca

            # Si la fila está completamente vacía en el nombre del equipo, saltarla
            if not val_equipo:
                continue

            # 2. Limpieza de datos (Quitar espacios y evitar strings "None")
            nombre = str(val_equipo).strip()
            operador = str(val_operador).strip() if val_operador else "SIN OPERADOR"
            ubicacion = str(val_planta).strip() if val_planta else None

            # Manejo de la columna F (Serie/Marca)
            # Como vienen juntos, los guardamos igual o puedes separarlos si tienen un guion
            serie = str(val_serie_m).strip() if val_serie_m else None
            marca = None  # Puedes dejarlo vacío o extraerlo si logras identificar un patrón

            # 3. Lógica para el Identificador (Determinar si viene en D o en E)
            # Si en la columna E viene algo que parece año o id, lo usamos.
            # Si E está vacío pero notas que el nombre en D ya incluye el ID (ej: "TRACTO RT2"),
            # podemos intentar extraerlo o dejar que compartan el valor.
            identificador = str(val_identif).strip() if val_identif else None

            # Si detectas que el identificador se puso en la columna D por error:
            # Ejemplo: si el valor de D es corto (ej: "RT2"), probablemente sea el identificador.
            if identificador is None and len(nombre) <= 6:
                identificador = nombre

            # 4. Guardar en Base de Datos
            # Usamos el nombre y el operador como llave única para actualizar o crear
            equipo_obj, created = Equipo.objects.update_or_create(
                nombre=nombre,
                operador=operador,
                defaults={
                    'identificador': identificador,
                    'serie': serie,
                    'marca': marca,
                    'ubicacion': ubicacion,
                }
            )

            if created:
                equipos_creados += 1
            else:
                equipos_actualizados += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"¡Catálogo de equipos cargado! Creados: {equipos_creados}, Actualizados: {equipos_actualizados}"
            )
        )
