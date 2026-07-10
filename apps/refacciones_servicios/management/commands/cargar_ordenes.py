import os
from datetime import datetime

from django.core.management import BaseCommand
from openpyxl.reader.excel import load_workbook

from apps.refacciones_servicios.models import DetalleOrdenCompra, Equipo, OrdenCompra, Proveedor


class Command(BaseCommand):
    help = 'Carga órdenes de compra desde un formato Excel no estructurado'

    def add_arguments(self, parser):
        parser.add_argument('excel_file', type=str, help='Ruta del archivo Excel (.xlsx)')

    def handle(self, *args, **options):
        ruta_archivo = options['excel_file']

        if not os.path.exists(ruta_archivo):
            self.stdout.write(self.style.ERROR(f"El archivo no existe: {ruta_archivo}"))
            return

        self.stdout.write(self.style.SUCCESS(f"Abriendo libro con openpyxl: {ruta_archivo}"))

        wb = load_workbook(ruta_archivo, data_only=True)
        todas_las_hojas = wb.sheetnames

        hojas_excluidas = ['obras', 'base datos']
        hojas_oc = [h for h in todas_las_hojas if h.strip().lower() not in hojas_excluidas]

        self.stdout.write(f'Número total de OC: {len(hojas_oc)}')

        for nombre_hoja in hojas_oc:
            ws = wb[nombre_hoja]
            self.stdout.write(f"Procesando hoja: {nombre_hoja}")

            # =========================================================================
            # 1. EXTRACCIÓN DE LA CABECERA
            # =========================================================================
            clave_oc = nombre_hoja.strip()

            fecha_cruda = ws['H5'].value
            nombre_prov_crudo = ws['C7'].value

            if isinstance(fecha_cruda, datetime):
                fecha_oc = fecha_cruda.date()
            elif isinstance(fecha_cruda, str):
                try:
                    fecha_oc = datetime.strptime(fecha_cruda.strip(), "%Y-%m-%d").date()
                except ValueError:
                    fecha_oc = datetime.now().date()
            else:
                fecha_oc = datetime.now().date()

            if nombre_prov_crudo:
                proveedor_obj = Proveedor.objects.filter(nombre__icontains=str(nombre_prov_crudo).strip()).first()
            else:
                proveedor_obj = None

            if not proveedor_obj:
                self.stdout.write(
                    self.style.WARNING(f"  -> Proveedor '{nombre_prov_crudo}' no encontrado. Saltando OC."))
                continue

            orden_obj, created = OrdenCompra.objects.update_or_create(
                clave=clave_oc,
                defaults={
                    'fecha': fecha_oc,
                    'proveedor': proveedor_obj
                }
            )

            if not created:
                DetalleOrdenCompra.objects.filter(orden=orden_obj).delete()

            # =========================================================================
            # 2. PROCESAMIENTO DE LAS PARTIDAS Y DETECCIÓN DE EQUIPOS (Hacia abajo)
            # =========================================================================
            fila_inicio = 13
            fila_fin = 24

            partidas_finales_a_guardar = []
            bloque_temporal_partidas = []
            partida_actual = None

            for r in range(fila_inicio, fila_fin + 1):
                val_desc = ws[f'B{r}'].value
                val_cantidad = ws[f'E{r}'].value
                val_unidad = ws[f'F{r}'].value
                val_precio = ws[f'G{r}'].value

                # Detectar el fin de la tabla
                if val_cantidad and "TOTAL" in str(val_cantidad).upper():
                    break
                if val_desc and "TOTAL" in str(val_desc).upper():
                    break

                texto_desc_str = str(val_desc).strip() if val_desc else ""

                # --- CASO ESPECIAL: Línea de cierre del Equipo ("Para RT3", "PARA OLLA 2", etc.) ---
                if val_desc and texto_desc_str.lower().startswith("para "):
                    # Primero cerramos la última partida que veníamos acumulando (si existe) y la pasamos al bloque
                    if partida_actual:
                        bloque_temporal_partidas.append(partida_actual)
                        partida_actual = None

                    # Extraer el nombre del equipo quitando la palabra "para "
                    nombre_equipo_detectado = texto_desc_str[5:].strip()  # Quita los primeros 5 caracteres "para "

                    # Asignamos este equipo a todas las partidas que estaban esperando en el bloque temporal
                    for p in bloque_temporal_partidas:
                        p['equipo_obj'] = nombre_equipo_detectado
                        partidas_finales_a_guardar.append(p)

                    # Vaciamos el bloque temporal para el siguiente grupo de equipos
                    bloque_temporal_partidas = []
                    continue

                # --- CASO 1: Inicio de una NUEVA partida (Tiene cantidad y unidad) ---
                if val_cantidad is not None and val_unidad is not None:
                    if partida_actual:
                        bloque_temporal_partidas.append(partida_actual)

                    partida_actual = {
                        'cantidad': int(val_cantidad),
                        'unidad': str(val_unidad).strip(),
                        'descripcion': texto_desc_str,
                        'precio': float(val_precio) if val_precio else 0.0,
                        'equipo_obj': ''  # Se resolverá cuando encontremos el "Para ..." abajo
                    }

                # --- CASO 2: Continuación de la descripción de la partida anterior ---
                elif val_desc and partida_actual:
                    partida_actual['descripcion'] += " " + texto_desc_str

            # Si salimos del bucle y quedó una partida colgada sin un "Para ..." explícito abajo
            if partida_actual:
                bloque_temporal_partidas.append(partida_actual)

            # =========================================================================
            # 3. GUARDADO DE DETALLES EN LA BASE DE DATOS
            # =========================================================================
            for p in partidas_finales_a_guardar:
                DetalleOrdenCompra.objects.create(
                    orden=orden_obj,
                    cantidad=p['cantidad'],
                    unidad=p['unidad'],
                    descripcion=p['descripcion'],
                    precio=p['precio'],
                    equipo=p['equipo_obj']
                )

        self.stdout.write(self.style.SUCCESS("¡Órdenes e historial de partidas cargados exitosamente!"))
