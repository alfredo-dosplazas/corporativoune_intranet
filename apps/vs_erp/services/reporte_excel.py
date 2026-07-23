import os
from itertools import groupby

import openpyxl
from django.conf import settings
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import Reference, PieChart
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from datetime import datetime

from apps.vs_erp.helpers import obtener_nombre_obra_descriptivo


# ----------------------------------------------------------------------
# 1. HOJA RESUMEN
# ----------------------------------------------------------------------
def construir_hoja_resumen(wb, reporte):
    ws = wb.active
    ws.title = "RESUMEN DE COMPRAS"
    ws.views.sheetView[0].showGridLines = True
    ws.views.sheetView[0].zoomScale = 85

    ruta_logo = os.path.join(settings.BASE_DIR, "static", "img", "logos", "LOGOS DOS PLAZAS.png")

    # -------------------------------------------------------------------
    # 0. AGREGAR LOGO DE LA EMPRESA
    # -------------------------------------------------------------------
    if ruta_logo and os.path.exists(ruta_logo):
        img = Image(ruta_logo)
        # Ajusta las dimensiones según las proporciones de tu imagen
        img.width = 120
        img.height = 120
        # Colocamos la imagen en la esquina superior izquierda
        ws.add_image(img, "B2")

    # Estilos
    font_titulo = Font(name="Calibri", size=14, bold=True, color="0F172A")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_normal = Font(name="Calibri", size=10)

    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_total_seccion = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    fill_total_general = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")
    fill_recuadro_hdr = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")

    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="0F172A")
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double)

    fmt_currency = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
    fmt_percent = '0.00%'

    # Encabezados generales
    primera_obra_id = reporte[0]['obra'] if reporte else ""
    nombre_obra_general = obtener_nombre_obra_descriptivo(primera_obra_id) if primera_obra_id else "GENERAL"

    ws["D2"] = "REPORTE RESUMEN DE COMPRAS"
    ws["D2"].font = font_titulo
    ws["D3"] = f"OBRA ENCABEZADO: {nombre_obra_general}"
    ws["D3"].font = font_bold
    ws["D4"] = f"FECHA DE EMISIÓN: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["D4"].font = font_normal

    headers_tabla = [
        "CASAS / N° CASAS", "PROTOTIPOS USADOS", "COMPRAS VS",
        "MONTO TOPE PTTO", "O.C. EFECTIVAMENTE COLOCADAS",
        "AVANCE DE OBRA %", "MAT. PENDIENTE DE COMPRAR & PTTO"
    ]

    start_row = 7
    for col_idx, text in enumerate(headers_tabla, start=2):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = start_row + 1
    filas_totales_secciones = []
    dict_totales_obras = {}

    for obra_item in reporte:
        id_obra = obra_item['obra']
        nombre_descriptivo = obtener_nombre_obra_descriptivo(id_obra)
        conceptos = obra_item['conceptos']

        id_upper = str(id_obra).upper().strip()
        debe_desglosar = id_upper.startswith("ED") or id_upper.startswith("URB")
        conceptos_n3 = [c for c in conceptos if c.get('NivelIdentacion') == 3] if debe_desglosar else []

        if debe_desglosar and conceptos_n3:
            fila_inicio_seccion = current_row
            for c in conceptos_n3:
                ws.cell(row=current_row, column=2, value=c.get('CantidadConcepto', 1)).alignment = Alignment(
                    horizontal="center")
                ws.cell(row=current_row, column=3, value=c.get('Prototipo', 'N/A')).alignment = Alignment(
                    horizontal="center")
                ws.cell(row=current_row, column=4, value=c.get('Concepto', ''))

                ws.cell(row=current_row, column=5,
                        value=float(c.get('PresupuestoMateriales', 0))).number_format = fmt_currency
                ws.cell(row=current_row, column=6,
                        value=float(c.get('EgresosMateriales', 0))).number_format = fmt_currency
                ws.cell(row=current_row, column=7,
                        value=f"=IF(E{current_row}>0, F{current_row}/E{current_row}, 0)").number_format = fmt_percent
                ws.cell(row=current_row, column=8, value=f"=E{current_row}-F{current_row}").number_format = fmt_currency

                for col in range(2, 9):
                    ws.cell(row=current_row, column=col).border = box_border
                current_row += 1

            fila_fin_seccion = current_row - 1
            ws.cell(row=current_row, column=2,
                    value=f"=SUM(B{fila_inicio_seccion}:B{fila_fin_seccion})").alignment = Alignment(
                horizontal="center")
            ws.cell(row=current_row, column=3, value="-").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=4, value=f"TOTAL {nombre_descriptivo}").font = font_bold
            ws.cell(row=current_row, column=5,
                    value=f"=SUM(E{fila_inicio_seccion}:E{fila_fin_seccion})").number_format = fmt_currency
            ws.cell(row=current_row, column=6,
                    value=f"=SUM(F{fila_inicio_seccion}:F{fila_fin_seccion})").number_format = fmt_currency
            ws.cell(row=current_row, column=7,
                    value=f"=IF(E{current_row}>0, F{current_row}/E{current_row}, 0)").number_format = fmt_percent
            ws.cell(row=current_row, column=8, value=f"=E{current_row}-F{current_row}").number_format = fmt_currency
        else:
            tot_pres = sum(float(c.get('PresupuestoMateriales', 0)) for c in conceptos if c.get('NivelIdentacion') == 1)
            tot_compras = sum(float(c.get('EgresosMateriales', 0)) for c in conceptos if c.get('NivelIdentacion') == 1)

            ws.cell(row=current_row, column=2, value="-").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=3, value="-").alignment = Alignment(horizontal="center")
            ws.cell(row=current_row, column=4, value=f"TOTAL {nombre_descriptivo}").font = font_bold
            ws.cell(row=current_row, column=5, value=tot_pres).number_format = fmt_currency
            ws.cell(row=current_row, column=6, value=tot_compras).number_format = fmt_currency
            ws.cell(row=current_row, column=7,
                    value=f"=IF(E{current_row}>0, F{current_row}/E{current_row}, 0)").number_format = fmt_percent
            ws.cell(row=current_row, column=8, value=f"=E{current_row}-F{current_row}").number_format = fmt_currency

        for col in range(2, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = fill_total_seccion
            cell.font = font_bold
            cell.border = total_border

        filas_totales_secciones.append(current_row)
        dict_totales_obras[nombre_descriptivo] = f"H{current_row}"
        current_row += 2

    # Total General
    fila_total_gen = current_row
    ws.cell(row=fila_total_gen, column=4, value="TOTAL VIVIENDA Y URBANIZACIÓN").font = font_bold
    sum_pres_str = "+".join([f"E{r}" for r in filas_totales_secciones])
    sum_compras_str = "+".join([f"F{r}" for r in filas_totales_secciones])

    ws.cell(row=fila_total_gen, column=5, value=f"={sum_pres_str}").number_format = fmt_currency
    ws.cell(row=fila_total_gen, column=6, value=f"={sum_compras_str}").number_format = fmt_currency
    ws.cell(row=fila_total_gen, column=7,
            value=f"=IF(E{fila_total_gen}>0, F{fila_total_gen}/E{fila_total_gen}, 0)").number_format = fmt_percent
    ws.cell(row=fila_total_gen, column=8, value=f"=E{fila_total_gen}-F{fila_total_gen}").number_format = fmt_currency

    for col in range(2, 9):
        cell = ws.cell(row=fila_total_gen, column=col)
        cell.fill = fill_total_general
        cell.font = font_bold
        cell.border = total_border

    # Recuadro Lateral
    row_recuadro = 7
    ws.cell(row=row_recuadro, column=10, value="CONCEPTO / OBRA").fill = fill_recuadro_hdr
    ws.cell(row=row_recuadro, column=10).font = font_header
    ws.cell(row=row_recuadro, column=11, value="PENDIENTE POR COMPRAR").fill = fill_recuadro_hdr
    ws.cell(row=row_recuadro, column=11).font = font_header

    start_chart_data_row = row_recuadro + 1
    row_recuadro += 1

    for nombre_desc, celda_ref in dict_totales_obras.items():
        c_lbl = ws.cell(row=row_recuadro, column=10, value=f"PENDIENTE {nombre_desc}")
        c_val = ws.cell(row=row_recuadro, column=11, value=f"={celda_ref}")
        c_lbl.border, c_val.border = box_border, box_border
        c_val.number_format = fmt_currency
        row_recuadro += 1

    end_chart_data_row = row_recuadro - 1

    c_lbl_tot = ws.cell(row=row_recuadro, column=10, value="TOTAL PENDIENTE POR COMPRAR")
    c_val_tot = ws.cell(row=row_recuadro, column=11, value=f"=SUM(K{start_chart_data_row}:K{end_chart_data_row})")
    c_lbl_tot.font, c_val_tot.font = font_bold, font_bold
    c_lbl_tot.fill, c_val_tot.fill = fill_total_seccion, fill_total_seccion
    c_lbl_tot.border, c_val_tot.border = total_border, total_border
    c_val_tot.number_format = fmt_currency

    # -------------------------------------------------------------------
    # 2. CONFIGURACIÓN DE LA GRÁFICA PIE (TOP-DOWN & SEPARADA)
    # -------------------------------------------------------------------
    pie = PieChart()  # Gráfica de pastel 2D plana (Top-Down view)
    pie.title = "Distribución del Material Pendiente por Comprar"

    data = Reference(ws, min_col=11, min_row=start_chart_data_row - 1, max_row=end_chart_data_row)
    labels = Reference(ws, min_col=10, min_row=start_chart_data_row, max_row=end_chart_data_row)

    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)

    # Separación/Espaciado entre rebanadas (Explode)
    # Define la distancia del centro hacia afuera (5 a 10 suele verse bastante bien)
    if pie.series:
        pie.series[0].explode = 8

        # Configuración de Etiquetas con Porcentaje (2 decimales)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = False  # Desactiva el monto absoluto
    pie.dataLabels.showPercent = True  # Muestra el porcentaje
    pie.dataLabels.numFmt = '0.00%'  # Formato exacto a 2 decimales

    pie.width, pie.height = 16, 11
    ws.add_chart(pie, f"J{row_recuadro + 3}")

    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['J'].width = 38
    ws.column_dimensions['K'].width = 25


def construir_hoja_volumen_precio(wb, obra_item):
    id_obra = obra_item.get('obra', '')
    nombre_desc = obtener_nombre_obra_descriptivo(id_obra)
    nombre_upper = nombre_desc.upper()

    if "EDIFICACI" not in nombre_upper and "URBA" not in nombre_upper:
        return

    sheet_title = f"VOL & PRECIO {nombre_desc}"[:29]
    ws = wb.create_sheet(title=sheet_title)

    ws.views.sheetView[0].showGridLines = True
    ws.views.sheetView[0].zoomScale = 85

    # -------------------------------------------------------------------
    # ESTILOS Y FORMATOS
    # -------------------------------------------------------------------
    font_titulo = Font(name="Calibri", size=14, bold=True, color="0F172A")
    font_subtitulo = Font(name="Calibri", size=10, bold=True, color="475569")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True, color="0F172A")
    font_normal = Font(name="Calibri", size=10, color="1E293B")

    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_total = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")
    fill_tarjeta = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="0F172A")

    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

    fmt_currency = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
    fmt_number = '#,##0.00;[Red](#,##0.00);"-"'

    # -------------------------------------------------------------------
    # 1. ENCABEZADO Y TARJETA DE TOTALES SUPERIOR (Filas 2 a 7)
    # -------------------------------------------------------------------
    ws["B2"] = f"ANÁLISIS DE VOLUMEN Y PRECIO - {nombre_desc}"
    ws["B2"].font = font_titulo
    ws["B3"] = "FAMILIA: CONCRETO"
    ws["B3"].font = font_subtitulo

    # Encabezado del Recuadro de Totales (Abarca de B5 a J5)
    ws.merge_cells("C5:K5")
    ws["C5"] = "RESUMEN GENERAL DE VARIACIONES DE CONCRETO"
    ws["C5"].font = font_header
    ws["C5"].fill = fill_header
    ws["C5"].alignment = Alignment(horizontal="center", vertical="center")

    # Etiquetas de Totales Superior (Fila 6)
    headers_totales = [
        ("C6", "VOLUMEN PRESUPUESTADO"),
        ("D6", "PRECIO PRESUPUESTADO"),
        ("E6", "VOLUMEN COMPRADO"),
        ("F6", "PRECIO COMPRADO"),
        ("G6", "DIF. PTTO - COMPRADO"),
        ("H6", "DIF. VOLUMEN Y PRECIO PTTO"),
        ("I6", "DIF. PRECIO Y COMPRA"),
        ("J6", "DIF. PENDIENTE COMPRO"),
        ("K6", "DIF. VOL PTTO VS COMPRADO")
    ]
    for cell_ref, text in headers_totales:
        c = ws[cell_ref]
        c.value = text
        c.font = font_bold
        c.fill = fill_tarjeta
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = box_border

    # -------------------------------------------------------------------
    # 2. TABLA DETALLADA DE MATERIALES
    # -------------------------------------------------------------------
    headers_tabla = [
        "ID INSUMO",
        "INSUMO / MATERIAL",
        "VOLUMEN PRESUPUESTADO",
        "PRECIO PRESUPUESTADO",
        "VOLUMEN COMPRADO",
        "PRECIO COMPRADO",
        "DIF. PTTO - COMPRADO",
        "DIF. VOLUMEN Y PRECIO PTTO",
        "DIF. PRECIO Y COMPRA",
        "DIF. PENDIENTE COMPRO",
        "DIF. VOL PTTO VS COMPRADO"
    ]

    start_row_tabla = 10
    for col_idx, text in enumerate(headers_tabla, start=1):  # Columnas A a K (1 a 11)
        cell = ws.cell(row=start_row_tabla, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = start_row_tabla + 1
    fila_inicio_data = current_row

    # Filtrar insumos pertenecientes a la familia CONCRETO
    materiales = obra_item.get('materiales', [])
    for m in materiales:
        CLAVE_FAMILIA_CONCRETO = 3
        clave_familia = int(m.get('IdFamilia', 0))

        if clave_familia == CLAVE_FAMILIA_CONCRETO:
            r = current_row

            # Datos básicos
            ws.cell(row=r, column=1, value=m.get('IdInsumo', '')).font = font_normal        # A: ID
            ws.cell(row=r, column=2, value=m.get('Material', m.get('Concepto', ''))).font = font_normal  # B: Material
            ws.cell(row=r, column=3, value=float(m.get('CantidadPresupuestada', 0))).number_format = fmt_number  # C: Vol Pres
            ws.cell(row=r, column=4, value=float(m.get('PresupuestoMateriales', 0))).number_format = fmt_currency  # D: Pcio Pres
            ws.cell(row=r, column=5, value=float(m.get('CantidadComprada', 0))).number_format = fmt_number  # E: Vol Comp
            ws.cell(row=r, column=6, value=float(m.get('EgresosMateriales', 0))).number_format = fmt_currency  # F: Pcio Comp

            # ---------------------------------------------------------------
            # Fórmulas de Fila (Columnas G a K)
            # ---------------------------------------------------------------
            # 1. G: Dif. Presupuesto - Comprado
            ws.cell(
                row=r, column=7,
                value=f"=D{r}-F{r}"
            ).number_format = fmt_currency

            # 2. H: Dif Vol y Pcio Ptto = (vol pres - vol comp) * (precio pres / vol pres)
            ws.cell(
                row=r, column=8,
                value=f"=IF(C{r}>0, (C{r}-E{r})*(D{r}/C{r}), 0)"
            ).number_format = fmt_currency

            # 3. I: DIF y PCIO = ((precio pres / vol pres) - (precio compr / vol compra)) * vol comprado
            ws.cell(
                row=r, column=9,
                value=f"=IF(AND(C{r}>0, E{r}>0), ((D{r}/C{r})-(F{r}/E{r}))*E{r}, 0)"
            ).number_format = fmt_currency

            # 4. J: Dif pte. compro = H + I
            ws.cell(
                row=r, column=10,
                value=f"=H{r}+I{r}"
            ).number_format = fmt_currency

            # 5. K: Dif. Volumen Presupuestado vs Comprado = C - E
            ws.cell(
                row=r, column=11,
                value=f"=C{r}-E{r}"
            ).number_format = fmt_number

            # Bordes a las 11 columnas (A a K)
            for col in range(1, 12):
                ws.cell(row=r, column=col).border = box_border

            current_row += 1

    fila_fin_data = current_row - 1

    # Si no se encontraron materiales de la familia Concreto
    if fila_fin_data < fila_inicio_data:
        ws.cell(row=current_row, column=1, value="No se encontraron insumos de la familia CONCRETO.").font = font_normal
        return

    # -------------------------------------------------------------------
    # 3. FILA DE TOTALES EN LA TABLA
    # -------------------------------------------------------------------
    fila_totales_tabla = current_row
    ws.cell(row=fila_totales_tabla, column=1, value="TOTAL CONCRETO").font = font_bold
    ws.cell(row=fila_totales_tabla, column=2, value="-").alignment = Alignment(horizontal="center")

    # Sumas por columna (de C a K)
    ws.cell(row=fila_totales_tabla, column=3, value=f"=SUM(C{fila_inicio_data}:C{fila_fin_data})").number_format = fmt_number    # Vol Pres
    ws.cell(row=fila_totales_tabla, column=4, value=f"=SUM(D{fila_inicio_data}:D{fila_fin_data})").number_format = fmt_currency  # Pcio Pres
    ws.cell(row=fila_totales_tabla, column=5, value=f"=SUM(E{fila_inicio_data}:E{fila_fin_data})").number_format = fmt_number    # Vol Comp
    ws.cell(row=fila_totales_tabla, column=6, value=f"=SUM(F{fila_inicio_data}:F{fila_fin_data})").number_format = fmt_currency  # Pcio Comp
    ws.cell(row=fila_totales_tabla, column=7, value=f"=SUM(G{fila_inicio_data}:G{fila_fin_data})").number_format = fmt_currency  # Dif Ptto - Comp
    ws.cell(row=fila_totales_tabla, column=8, value=f"=SUM(H{fila_inicio_data}:H{fila_fin_data})").number_format = fmt_currency  # Dif Vol & Pcio
    ws.cell(row=fila_totales_tabla, column=9, value=f"=SUM(I{fila_inicio_data}:I{fila_fin_data})").number_format = fmt_currency  # Dif Pcio & Comp
    ws.cell(row=fila_totales_tabla, column=10, value=f"=SUM(J{fila_inicio_data}:J{fila_fin_data})").number_format = fmt_currency # Dif Pte Compro
    ws.cell(row=fila_totales_tabla, column=11, value=f"=SUM(K{fila_inicio_data}:K{fila_fin_data})").number_format = fmt_number   # Dif Vol Ptto vs Comp

    for col in range(1, 12):
        cell = ws.cell(row=fila_totales_tabla, column=col)
        cell.fill = fill_total
        cell.font = font_bold
        cell.border = total_border

    # -------------------------------------------------------------------
    # 4. LLENAR LA TARJETA DE TOTALES SUPERIOR (Fila 7: B a J)
    # -------------------------------------------------------------------
    ws["C7"] = f"=C{fila_totales_tabla}"  # Vol Pres Total
    ws["C7"].number_format = fmt_number

    ws["D7"] = f"=D{fila_totales_tabla}"  # Pcio Pres Total
    ws["D7"].number_format = fmt_currency

    ws["E7"] = f"=E{fila_totales_tabla}"  # Vol Comp Total
    ws["E7"].number_format = fmt_number

    ws["F7"] = f"=F{fila_totales_tabla}"  # Pcio Comp Total
    ws["F7"].number_format = fmt_currency

    ws["G7"] = f"=G{fila_totales_tabla}"  # Dif Ptto - Comp Total
    ws["G7"].number_format = fmt_currency

    ws["H7"] = f"=H{fila_totales_tabla}"  # Dif Vol & Pcio Total
    ws["H7"].number_format = fmt_currency

    ws["I7"] = f"=I{fila_totales_tabla}"  # Dif Pcio & Comp Total
    ws["I7"].number_format = fmt_currency

    ws["J7"] = f"=J{fila_totales_tabla}"  # Dif Pte Compro Total
    ws["J7"].number_format = fmt_currency

    ws["K7"] = f"=K{fila_totales_tabla}"  # Dif Vol Ptto vs Comp Total
    ws["K7"].number_format = fmt_number

    for col in ["C7", "D7", "E7", "F7", "G7", "H7", "I7", "J7", "K7"]:
        c = ws[col]
        c.font = font_bold
        c.fill = fill_tarjeta
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = box_border

    # -------------------------------------------------------------------
    # 5. ANCHO DE COLUMNAS
    # -------------------------------------------------------------------
    ws.column_dimensions['A'].width = 18  # ID Insumo
    ws.column_dimensions['B'].width = 45  # Insumo / Material
    ws.column_dimensions['C'].width = 24  # Vol Pres
    ws.column_dimensions['D'].width = 24  # Pcio Pres
    ws.column_dimensions['E'].width = 24  # Vol Comp
    ws.column_dimensions['F'].width = 24  # Pcio Comp
    ws.column_dimensions['G'].width = 26  # Dif Ptto - Comp
    ws.column_dimensions['H'].width = 28  # Dif Vol y Pcio Ptto
    ws.column_dimensions['I'].width = 28  # Dif Pcio y Compra
    ws.column_dimensions['J'].width = 28  # Dif Pte Compro
    ws.column_dimensions['K'].width = 28  # Dif Vol Ptto vs Comp


def construir_hoja_retenciones(wb, obra_item):
    retenciones = obra_item.get('retenciones', [])

    # Solo construimos la hoja si existen retenciones para esta obra
    if not retenciones:
        return

    id_obra = obra_item.get('obra', '')
    nombre_desc = obtener_nombre_obra_descriptivo(id_obra)

    # openpyxl limita el título de la pestaña a 31 caracteres
    sheet_title = f"OC RETENCIONES {nombre_desc}"[:29]
    ws = wb.create_sheet(title=sheet_title)

    # ----------------------------------------------------
    # CONFIGURACIÓN DE VISTA (Zoom 85% y Cuadrícula)
    # ----------------------------------------------------
    ws.views.sheetView[0].showGridLines = True

    # ----------------------------------------------------
    # ESTILOS CORPORATIVOS
    # ----------------------------------------------------
    font_titulo = Font(name="Calibri", size=14, bold=True, color="0F172A")
    font_subtitulo = Font(name="Calibri", size=10, bold=True, color="475569")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_normal = Font(name="Calibri", size=10, color="1E293B")

    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_total = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")

    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="0F172A")

    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

    fmt_currency = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'

    # ----------------------------------------------------
    # ENCABEZADO DE LA HOJA
    # ----------------------------------------------------
    ws["A1"] = "REPORTE DE RETENCIONES EN ÓRDENES DE COMPRA"
    ws["A1"].font = font_titulo

    ws["A2"] = f"OBRA: {nombre_desc} | {id_obra}"
    ws["A2"].font = font_subtitulo

    # ----------------------------------------------------
    # TABLA DE RETENCIONES
    # ----------------------------------------------------
    headers = [
        "ID INSUMO",
        "ID ORDEN DE COMPRA",
        "CLAVE / ID OBRA",
        "IMPORTE RETENCIÓN"
    ]

    start_row = 4
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = start_row + 1

    for item in retenciones:
        # Columna A: ID Insumo
        ws.cell(row=current_row, column=1, value=item.get('IdInsumo', '')).alignment = Alignment(horizontal="center")

        # Columna B: ID Orden de Compra
        ws.cell(row=current_row, column=2, value=item.get('IdOrdenCompra', '')).alignment = Alignment(
            horizontal="center")

        # Columna C: ID Obra
        ws.cell(row=current_row, column=3, value=id_obra).alignment = Alignment(horizontal="center")

        # Columna D: Importe de la Retención
        c_imp = ws.cell(row=current_row, column=4, value=float(item.get('importeR', 0)))
        c_imp.number_format = fmt_currency
        c_imp.alignment = Alignment(horizontal="right")

        # Aplicar bordes y fuentes
        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.font = font_normal
            cell.border = box_border

        current_row += 1

    # ----------------------------------------------------
    # FILA DE TOTALES
    # ----------------------------------------------------
    fila_fin_datos = current_row - 1

    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=2, value="")
    ws.cell(row=current_row, column=3, value="TOTAL RETENCIONES").font = font_bold
    ws.cell(row=current_row, column=3).alignment = Alignment(horizontal="right")

    # Fórmula de Suma
    if fila_fin_datos >= start_row + 1:
        c_tot = ws.cell(row=current_row, column=4, value=f"=SUM(D{start_row + 1}:D{fila_fin_datos})")
    else:
        c_tot = ws.cell(row=current_row, column=4, value=0)

    c_tot.number_format = fmt_currency
    c_tot.font = font_bold

    # Estilos para la fila final
    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = fill_total
        cell.border = total_border

    # ----------------------------------------------------
    # ANCHOS DE COLUMNA
    # ----------------------------------------------------
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 25


def construir_hoja_compras_vs(wb, obra_item):
    id_obra = obra_item.get('obra', '')
    nombre_desc = obtener_nombre_obra_descriptivo(id_obra)

    sheet_title = f"COMPRAS {nombre_desc} VS"[:29]
    ws = wb.create_sheet(title=sheet_title)

    ws.views.sheetView[0].showGridLines = True

    # ----------------------------------------------------
    # ESTILOS
    # ----------------------------------------------------
    font_titulo = Font(name="Calibri", size=14, bold=True, color="0F172A")
    font_subtitulo = Font(name="Calibri", size=10, bold=True, color="475569")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    # Fuentes y rellenos según Jerarquía / Nivel
    font_nivel_1 = Font(name="Calibri", size=10, bold=True, color="0F172A")
    font_nivel_2 = Font(name="Calibri", size=10, bold=True, color="334155")
    font_normal = Font(name="Calibri", size=10, color="1E293B")

    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_nivel_1 = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")  # Gris muy suave
    fill_total = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")

    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="0F172A")

    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

    fmt_currency = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
    fmt_number = '#,##0.00;[Red]-#,##0.00;"-"'

    # Encabezado
    ws["A1"] = "DETALLE DE CONCEPTOS VS COMPRAS"
    ws["A1"].font = font_titulo
    ws["A2"] = f"OBRA: {nombre_desc} | {id_obra}"
    ws["A2"].font = font_subtitulo

    # Encabezados Tabla
    headers = [
        "ID CONCEPTO", "CLAVE CONCEPTO", "DESCRIPCIÓN",
        "CANTIDAD", "COSTO PRESUPUESTADO", "IMPORTE PRESUPUESTADO", "EGRESOS (O.C.)", "DIFERENCIA (IMP. PTTO - EGRESOS)"
    ]

    start_row = 4
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = start_row + 1
    conceptos = obra_item.get('conceptos', [])

    for c in conceptos:
        nivel = int(c.get('NivelIdentacion', 1))

        # 1. Agrupamiento de filas (Acordeón nativo de Excel)
        # Nivel 1 = outline_level 0, Nivel 2 = outline_level 1, etc.
        if nivel > 1:
            ws.row_dimensions[current_row].outline_level = nivel - 1

        ws.cell(row=current_row, column=1, value=c.get('IdConceptoObra', '')).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=2, value=c.get('ClaveConceptoObra', '')).alignment = Alignment(
            horizontal="center")

        # 2. Sangría Visual en la descripción (indent)
        # Sangría de nivel: Nivel 1 (0), Nivel 2 (2 de sangría), Nivel 3 (4 de sangría)
        indent_level = max(0, (nivel - 1) * 2)
        cell_desc = ws.cell(row=current_row, column=3, value=c.get('Concepto', c.get('Descripcion', '')))
        cell_desc.alignment = Alignment(horizontal="left", indent=indent_level)

        # Valores
        costo_presupuestado = float(c.get('PresupuestoMateriales', 0)) / float(c.get('CantidadConcepto', 0))

        ws.cell(row=current_row, column=4, value=float(c.get('CantidadConcepto', 0))).number_format = fmt_number
        ws.cell(row=current_row, column=5, value=float(costo_presupuestado)).number_format = fmt_currency
        ws.cell(row=current_row, column=6, value=float(c.get('PresupuestoMateriales', 0))).number_format = fmt_currency
        ws.cell(row=current_row, column=7, value=float(c.get('EgresosMateriales', 0))).number_format = fmt_currency

        # Fórmula de diferencia
        ws.cell(row=current_row, column=8, value=f"=F{current_row}-G{current_row}").number_format = fmt_currency

        # 3. Estilos dinámicos según el Nivel de Identación
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = box_border

            if nivel == 1:
                cell.font = font_nivel_1
                cell.fill = fill_nivel_1
            elif nivel == 2:
                cell.font = font_nivel_2
            else:
                cell.font = font_normal

        current_row += 1

    # ----------------------------------------------------
    # FILA DE TOTALES
    # ----------------------------------------------------
    fila_fin_datos = current_row - 1

    # ws.cell(row=current_row, column=3, value="TOTALES").font = font_nivel_1
    #
    # if fila_fin_datos >= start_row + 1:
    #     # Sumamos solo los elementos de Nivel 1 si tu data viene jerarquizada para evitar duplicar totales,
    #     # o hacemos SUMA normal si la data solo trae los conceptos hoja.
    #     ws.cell(row=current_row, column=5, value=f"=SUM(E{start_row + 1}:E{fila_fin_datos})").number_format = fmt_number
    #     ws.cell(row=current_row, column=6,
    #             value=f"=SUM(F{start_row + 1}:F{fila_fin_datos})").number_format = fmt_currency
    #     ws.cell(row=current_row, column=7,
    #             value=f"=SUM(G{start_row + 1}:G{fila_fin_datos})").number_format = fmt_currency
    #     ws.cell(row=current_row, column=8, value=f"=F{current_row}-G{current_row}").number_format = fmt_currency

    for col in range(1, 9):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = fill_total
        cell.font = font_nivel_1
        cell.border = total_border

    # Anchos de columna
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 80  # Espacio amplio para la sangría de texto
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 30


def construir_hoja_compras_familia(wb, obra_item):
    familias = obra_item.get('familias', [])
    id_obra = obra_item.get('obra', '')
    nombre_desc = obtener_nombre_obra_descriptivo(id_obra)

    sheet_title = f"COMPRAS {nombre_desc} FAMILIA"[:29]
    ws = wb.create_sheet(title=sheet_title)

    # ----------------------------------------------------
    # CONFIGURACIÓN DE VISTA Y CONGELADO DE PANELS
    # ----------------------------------------------------
    ws.views.sheetView[0].showGridLines = True

    # Congelar de la fila 5 hacia arriba (Fila 1 a 5 quedan fijas al hacer scroll)
    ws.freeze_panes = "A6"

    # ----------------------------------------------------
    # ESTILOS CORPORATIVOS
    # ----------------------------------------------------
    font_titulo = Font(name="Calibri", size=14, bold=True, color="0F172A")
    font_subtitulo = Font(name="Calibri", size=10, bold=True, color="475569")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_normal = Font(name="Calibri", size=10, color="1E293B")

    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_total = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")

    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="0F172A")

    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

    fmt_currency = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'

    # ----------------------------------------------------
    # ENCABEZADO DE LA HOJA
    # ----------------------------------------------------
    ws["A1"] = "RESUMEN DE COMPRAS POR FAMILIA DE INSUMOS"
    ws["A1"].font = font_titulo

    ws["A2"] = f"OBRA: {nombre_desc} | {id_obra}"
    ws["A2"].font = font_subtitulo

    # ----------------------------------------------------
    # FILA 4: TOTALES GENERALES (FIJA ARRIBA)
    # ----------------------------------------------------
    row_totales = 4
    num_items = len(familias)

    # Filas donde vivirán los datos
    row_datos_inicio = 6
    row_datos_fin = row_datos_inicio + num_items - 1

    ws.cell(row=row_totales, column=1, value="")
    ws.cell(row=row_totales, column=2, value="TOTAL GENERAL").font = font_bold

    if num_items > 0:
        ws.cell(row=row_totales, column=3,
                value=f"=SUM(C{row_datos_inicio}:C{row_datos_fin})").number_format = fmt_currency
        ws.cell(row=row_totales, column=4,
                value=f"=SUM(D{row_datos_inicio}:D{row_datos_fin})").number_format = fmt_currency
        ws.cell(row=row_totales, column=5,
                value=f"=C{row_totales}-D{row_totales}").number_format = fmt_currency
    else:
        for col_idx in range(3, 6):
            ws.cell(row=row_totales, column=col_idx, value=0).number_format = fmt_currency

    # Estilos de la fila de totales
    for col in range(1, 6):
        cell = ws.cell(row=row_totales, column=col)
        cell.fill = fill_total
        cell.font = font_bold
        cell.border = total_border

    # ----------------------------------------------------
    # FILA 5: ENCABEZADOS DE TABLA
    # ----------------------------------------------------
    headers = [
        "ID FAMILIA",
        "FAMILIA / GRUPO DE INSUMOS",
        "PRESUPUESTO MATERIALES",
        "EGRESOS (O.C. COLOCADAS)",
        "DIFERENCIA (PTTO - EGRESOS)"
    ]

    row_headers = 5
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row_headers, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ----------------------------------------------------
    # LLENADO DE DATOS (DESDE LA FILA 6)
    # ----------------------------------------------------
    current_row = row_datos_inicio

    for f in familias:
        # Columna A: ID Familia
        ws.cell(row=current_row, column=1,
                value=f.get('IdFamilia', f.get('IdFamiliaInsumos', ''))).alignment = Alignment(horizontal="center")

        # Columna B: Descripción de la Familia
        ws.cell(row=current_row, column=2, value=f.get('Familia', f.get('Descripcion', '')))

        # Columna C: Presupuesto
        c_pres = ws.cell(row=current_row, column=3,
                         value=float(f.get('PresupuestoMateriales', f.get('ImportePresupuestado', 0))))
        c_pres.number_format = fmt_currency

        # Columna D: Egresos
        c_egre = ws.cell(row=current_row, column=4,
                         value=float(f.get('EgresosMateriales', f.get('ImporteEgresos', 0))))
        c_egre.number_format = fmt_currency

        # Columna E: Diferencia
        c_dif = ws.cell(row=current_row, column=5, value=f"=C{current_row}-D{current_row}")
        c_dif.number_format = fmt_currency

        # Aplicar bordes y fuentes
        for col in range(1, 6):
            cell = ws.cell(row=current_row, column=col)
            cell.font = font_normal
            cell.border = box_border

        current_row += 1

    # ----------------------------------------------------
    # ANCHOS DE COLUMNA
    # ----------------------------------------------------
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25


def construir_hoja_compras_material(wb, obra_item):
    lista_materiales = obra_item.get('materiales', [])
    id_obra = obra_item.get('obra', '')
    nombre_desc = obtener_nombre_obra_descriptivo(id_obra)

    sheet_title = f"COMPRAS {nombre_desc} MATERIAL"[:29]
    ws = wb.create_sheet(title=sheet_title)

    # ----------------------------------------------------
    # CONFIGURACIÓN DE VISTA Y FREEZE PANES
    # ----------------------------------------------------
    ws.views.sheetView[0].showGridLines = True

    # Congelar de la fila 5 hacia arriba (Filas 1 a 5 quedan fijas al hacer scroll)
    ws.freeze_panes = "A6"

    # ----------------------------------------------------
    # ESTILOS CORPORATIVOS
    # ----------------------------------------------------
    font_titulo = Font(name="Calibri", size=14, bold=True, color="0F172A")
    font_subtitulo = Font(name="Calibri", size=10, bold=True, color="475569")
    font_header = Font(name="Calibri", size=10, bold=True, color="FFFFFF")

    font_familia = Font(name="Calibri", size=10, bold=True, color="0F172A")
    font_normal = Font(name="Calibri", size=10, color="1E293B")
    font_bold = Font(name="Calibri", size=10, bold=True)

    fill_header = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    fill_familia = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_total = PatternFill(start_color="CBD5E1", end_color="CBD5E1", fill_type="solid")

    border_thin = Side(border_style="thin", color="CBD5E1")
    border_double = Side(border_style="double", color="0F172A")

    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    total_border = Border(top=border_thin, bottom=border_double, left=border_thin, right=border_thin)

    fmt_currency = '"$"#,##0.00;[Red]("$"#,##0.00);"-"'
    fmt_number = '#,##0.00;[Red]-#,##0.00;"-"'

    # ----------------------------------------------------
    # ENCABEZADO DE LA HOJA
    # ----------------------------------------------------
    ws["A1"] = "DETALLE DE COMPRAS POR MATERIAL E INSUMO"
    ws["A1"].font = font_titulo

    ws["A2"] = f"OBRA: {nombre_desc} | {id_obra}"
    ws["A2"].font = font_subtitulo

    # ----------------------------------------------------
    # FILA 4: TOTAL GENERAL (RESERVAMOS LA POSICIÓN ARRIBA)
    # ----------------------------------------------------
    row_totales = 4
    ws.cell(row=row_totales, column=1, value="")
    ws.cell(row=row_totales, column=2, value="TOTAL GENERAL DE MATERIALES").font = font_bold

    # ----------------------------------------------------
    # FILA 5: ENCABEZADOS DE LA TABLA
    # ----------------------------------------------------
    headers = [
        "CLAVE / INSUMO",
        "DESCRIPCIÓN DEL MATERIAL",
        "UNIDAD",
        "CANT. PTTO",
        "PRECIO PTTO",
        "IMPORTE PTTO",
        "CANT. COMPRADA",
        "PRECIO COMPRA",
        "IMPORTE COMPRA",
        "DIFERENCIA (PTTO - EGRESOS)"
    ]

    row_headers = 5
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row_headers, column=col_idx, value=text)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ----------------------------------------------------
    # LLENADO DE DATOS (DESDE LA FILA 6)
    # ----------------------------------------------------
    current_row = 6
    filas_familias_padre = []  # Guardaremos las filas de los renglones PADRE (Familias)

    for nombre_familia, grupo_materiales in groupby(lista_materiales, key=lambda x: x.get('Familia', 'SIN FAMILIA')):
        items_materiales = list(grupo_materiales)

        # 1. RENGLÓN PADRE (FAMILIA)
        row_familia_inicio = current_row
        filas_familias_padre.append(row_familia_inicio)

        id_fam_repr = items_materiales[0].get('IdFamilia', '') if items_materiales else ''

        ws.cell(row=current_row, column=1, value=id_fam_repr).alignment = Alignment(horizontal="center")
        ws.cell(row=current_row, column=2, value=f"FAMILIA: {nombre_familia}").alignment = Alignment(horizontal="left")

        current_row += 1
        row_hijos_inicio = current_row

        # 2. RENGLONES HIJOS (MATERIALES)
        for mat in items_materiales:
            ws.row_dimensions[current_row].outline_level = 1

            # Clave Insumo
            ws.cell(row=current_row, column=1, value=mat.get('IdInsumo', '')).alignment = Alignment(horizontal="center")

            # Descripción con sangría
            c_desc = ws.cell(row=current_row, column=2, value=mat.get('Material', ''))
            c_desc.alignment = Alignment(horizontal="left", indent=2)

            # Unidad
            ws.cell(row=current_row, column=3, value=mat.get('UnidadInsumo', '')).alignment = Alignment(
                horizontal="center")

            # PRESUPUESTO
            cant_ptto = float(mat.get('CantidadPresupuestada', 0.0))
            imp_ptto = float(mat.get('PresupuestoMateriales', 0.0))
            precio_ptto = (imp_ptto / cant_ptto) if cant_ptto > 0 else 0.0

            ws.cell(row=current_row, column=4, value=cant_ptto).number_format = fmt_number
            ws.cell(row=current_row, column=5, value=precio_ptto).number_format = fmt_currency
            ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}").number_format = fmt_currency

            # COMPRAS / EGRESOS
            cant_comp = float(mat.get('CantidadComprada', 0.0))
            imp_comp = float(mat.get('EgresosMateriales', 0.0))
            precio_comp = (imp_comp / cant_comp) if cant_comp > 0 else 0.0

            ws.cell(row=current_row, column=7, value=cant_comp).number_format = fmt_number
            ws.cell(row=current_row, column=8, value=precio_comp).number_format = fmt_currency
            ws.cell(row=current_row, column=9, value=f"=G{current_row}*H{current_row}").number_format = fmt_currency

            # DIFERENCIA (Importe Ptto - Importe Compra)
            ws.cell(row=current_row, column=10, value=f"=F{current_row}-I{current_row}").number_format = fmt_currency

            # Estilos de fila hijo
            for col in range(1, 11):
                cell = ws.cell(row=current_row, column=col)
                cell.font = font_normal
                cell.border = box_border

            current_row += 1

        row_hijos_fin = current_row - 1

        # 3. SUMAS Y FÓRMULAS EN EL PADRE (FAMILIA)
        if row_hijos_fin >= row_hijos_inicio:
            ws.cell(row=row_familia_inicio, column=6,
                    value=f"=SUM(F{row_hijos_inicio}:F{row_hijos_fin})").number_format = fmt_currency
            ws.cell(row=row_familia_inicio, column=9,
                    value=f"=SUM(I{row_hijos_inicio}:I{row_hijos_fin})").number_format = fmt_currency
            ws.cell(row=row_familia_inicio, column=10,
                    value=f"=F{row_familia_inicio}-I{row_familia_inicio}").number_format = fmt_currency
        else:
            ws.cell(row=row_familia_inicio, column=6, value=0.0).number_format = fmt_currency
            ws.cell(row=row_familia_inicio, column=9, value=0.0).number_format = fmt_currency
            ws.cell(row=row_familia_inicio, column=10, value=0.0).number_format = fmt_currency

        # Estilos de fila padre (Familia)
        for col in range(1, 11):
            cell = ws.cell(row=row_familia_inicio, column=col)
            cell.fill = fill_familia
            cell.font = font_familia
            cell.border = box_border

    # ----------------------------------------------------
    # APLICAR FÓRMULAS Y ESTILOS A LA FILA 4 (TOTAL GENERAL)
    # ----------------------------------------------------
    if filas_familias_padre:
        # Sumamos los Totales de cada Familia (Padre)
        str_sum_ptto = "+".join([f"F{r}" for r in filas_familias_padre])
        str_sum_egre = "+".join([f"I{r}" for r in filas_familias_padre])

        ws.cell(row=row_totales, column=6, value=f"={str_sum_ptto}").number_format = fmt_currency
        ws.cell(row=row_totales, column=9, value=f"={str_sum_egre}").number_format = fmt_currency
        ws.cell(row=row_totales, column=10, value=f"=F{row_totales}-I{row_totales}").number_format = fmt_currency
    else:
        ws.cell(row=row_totales, column=6, value=0.0).number_format = fmt_currency
        ws.cell(row=row_totales, column=9, value=0.0).number_format = fmt_currency
        ws.cell(row=row_totales, column=10, value=0.0).number_format = fmt_currency

    # Estilos de la fila total (arriba)
    for col in range(1, 11):
        cell = ws.cell(row=row_totales, column=col)
        cell.fill = fill_total
        cell.font = font_bold
        cell.border = total_border

    # ----------------------------------------------------
    # ANCHOS DE COLUMNA
    # ----------------------------------------------------
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 22


def generar_excel_reporte_completo(reporte):
    wb = openpyxl.Workbook()

    construir_hoja_resumen(wb, reporte)

    for obra_item in reporte:
        construir_hoja_retenciones(wb, obra_item)
        construir_hoja_volumen_precio(wb, obra_item)
        construir_hoja_compras_vs(wb, obra_item)
        construir_hoja_compras_familia(wb, obra_item)
        construir_hoja_compras_material(wb, obra_item)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="Reporte_Presupuestos_Compras.xlsx"'
    wb.save(response)
    return response
